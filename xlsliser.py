import openpyxl
from openpyxl import Workbook
import re


def svod():
    workFile = openpyxl.load_workbook("D:/excel/test/Врачи.xlsx")
    sheet = workFile.active
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def main():
    result = {}
    lists = list(svod())
    for lst in lists:
        if lst[6] not in result:
            result[lst[6]] = []
        result[lst[6]].append(lst)
    for mo in result:

        print("-------------" + mo + "-----------------")
        wbMo = Workbook()
        wsMo = wbMo.active
        row = 1
        col = 1
        for i in result[mo]:
            for j in i:
                wsMo.cell(row=1, column=col).value = lists[0][col-1]
                wsMo.cell(row=row+1, column=col).value = j
                col += 1
            col = 1
            row += 1
            print(i)
        cleanName = re.sub(r'[^а-яА-Я0-9№ ]+', '', mo)
        wbMo.save("D:/excel/test/1/" + cleanName + '.xlsx')


if __name__ == '__main__':
    main()
